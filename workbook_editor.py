# workbook_editor.py
# Minimal, dependency: openpyxl
# Intended for use in Pyodide or any Python environment where openpyxl is available.

from typing import Tuple, Optional
import io
from openpyxl import load_workbook


def process_workbook(wb_bytes: bytes, b2_value) -> Tuple[Optional[object], bytes]:
    """
    Write b2_value to cell B2 of the workbook given by wb_bytes,
    read the value stored in cell C2 (cached value if it's a formula),
    and return (c2_value, edited_workbook_bytes).

    Args:
      wb_bytes: bytes of an .xlsx file.
      b2_value: value to write into B2 (string, number, etc).

    Returns:
      (c2_value, out_bytes)
        c2_value: the value read from C2 (could be None or any Excel-stored type).
        out_bytes: bytes of the edited workbook (xlsx).
    """
    # Load workbook from bytes
    stream = io.BytesIO(wb_bytes)
    wb = load_workbook(stream, data_only=False)
    ws = wb.active

    # Write to B2
    ws["B2"] = b2_value

    # Read C2 (cached value stored in the file)
    try:
        c2_val = ws["C2"].value
    except Exception as e:
        c2_val = f"READ_ERROR: {e}"

    # Save modified workbook to bytes
    out_stream = io.BytesIO()
    wb.save(out_stream)
    out_bytes = out_stream.getvalue()

    return c2_val, out_bytes


# Optional: allow running locally for quick tests
if __name__ == "__main__":
    import sys
    from pathlib import Path

    if len(sys.argv) < 4:
        print("Usage: python workbook_editor.py <input_xlsx> <b2_value> <output_xlsx>")
        sys.exit(1)

    in_path = Path(sys.argv[1])
    b2_val = sys.argv[2]
    out_path = Path(sys.argv[3])

    if not in_path.exists():
        print(f"Input file not found: {in_path}")
        sys.exit(2)

    with open(in_path, "rb") as f:
        wb_bytes_local = f.read()

    c2, out_bytes = process_workbook(wb_bytes_local, b2_val)
    with open(out_path, "wb") as f:
        f.write(out_bytes)

    print(f"Wrote B2 = {b2_val}; read C2 = {c2}; saved edited workbook to {out_path}")
